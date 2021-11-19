from openpyxl import load_workbook
from pprint import pprint

wb = load_workbook(r"data.xlsm")
ws = wb["вопросы"]

quests_list = []
answer_list = []
for row in ws.iter_rows(min_row=3, max_col=7, max_row=129):
    quests = {}
    for cell in row:
        if cell.coordinate[0] != 'A' and cell.coordinate[0] != 'B':
            if cell.coordinate[0] == 'C':
                quests[cell.value] = []
    quests_list.append(quests)

for row in ws.iter_rows(min_row=3, max_col=7, max_row=129):
    for cell in row:
        print(f"coordinate: {cell.coordinate}\tvalue: {cell.value}\tcolor: {cell.fill.start_color.index}")
        if cell.coordinate[0] != 'A' and cell.coordinate[0] != 'B':
            for ans_dict in quests_list:
                for k, v in ans_dict.items():
                    if k == cell.value:
                        for ce in row:
                            if ce.coordinate[0] != 'A' and ce.coordinate[0] != 'B' and ce.coordinate[0] != 'C':
                                if ce.fill.start_color.index == 'FF92D050':
                                    v.append((True, ce.value))
                                else:
                                    v.append((False, ce.value))

# print(quests_list)

for quest_dict in quests_list:
    with open("data1.xml", "a+") as file:
        file.write('<my:question>')
    answ_l = []
    for k, v in quest_dict.items():

        str_quest = f"""
    <my:qtext><div style="FONT-FAMILY: Microsoft Sans Serif" align="center" xmlns="http://www.w3.org/1999/xhtml" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><strong><font face="Calibri">{k}</font></strong></div></my:qtext>
        """

        with open("data1.xml", "a+") as file:
            file.write(str_quest)

        for anws in v:
            str_answ = f"""
    <my:answer>
        <my:astatus>{'Правильный ответ' if anws[0] is True else 'Неправильный ответ'}</my:astatus>
        <my:atext>{anws[1]}</my:atext>
    </my:answer>
"""
            with open("data1.xml", "a+") as file:
                file.write(str_answ)
        with open("data1.xml", "a+") as file:
            file.write('<my:qhelp></my:qhelp>\n</my:question>')
